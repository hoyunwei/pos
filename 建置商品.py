#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
從主檔「三支筆商品營運管理.xlsx →文博商品清單」建置 POS 匯入檔 + index.html 用的 seed。

規則（2026-07-31 小何確認）：
  - 分類名沿用主檔（明信片 / A4海報 / 細長海報 / 貼紙 / 造型卡 / 書籤 / L夾 /
    冰箱貼 / 冰箱貼-春節 / 帆布袋 / 水晶貼 / 凸版明信片 / 凸版方形）
  - 庫存＝「帶去數量」，空白→0（收銀台灰掉不可點）
  - 缺價 fallback：造型卡 60、L夾 100、凸版明信片 80、凸版方形 250、帆布袋 450
    書籤 = 贈品，主檔定價欄寫「贈品」→ 這裡填「加購價 29」（滿 $300 送 1 張，也可 $29 加購）
    2026-07-31 更新：主檔已補齊定價，多數 fallback 不會觸發，留著只是保險
  - 冰箱貼-春節 定價 0/空 → RH 系列補 160、其餘補 150
  - 水晶貼 23 款共用 CS2601 → 合成 3 個尺寸 SKU：小片$35 / 中片$70 / 大片$120，
    庫存＝盤點表各尺寸帶去量加總
產出：商品匯入_文博.json（含圖 base64）＋ seed_snippet.txt（貼進 index.html）

⚠️ 只要改了 seed（定價／商品／分類），index.html 的 PATCH_VER 一定要 +1，
   否則已經用過的裝置不會吃到變更。
⚠️ 水晶貼刻意不附圖（SKIP_IMG_IDS）：CS2601_水晶貼.jpg 是程式產的佔位圖，
   上面紅字寫「未定價」，但三個尺寸早已定價 35／70／120，擺在收銀台會誤導客人。
   小何 2026-07-31 決定這三張卡不要圖。之後有真照片再從這裡拿掉即可。
"""
import openpyxl, json, os, base64, io
from PIL import Image, ImageOps

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "..", "..", "商品管理", "三支筆商品營運管理.xlsx")
IMG_DIR = os.path.join(HERE, "商品圖片")
OUT_JSON = os.path.join(HERE, "商品匯入_文博.json")
OUT_SEED = os.path.join(HERE, "seed_snippet.txt")
MAX_SIDE, QUALITY = 400, 70
EXTS = (".jpg", ".jpeg", ".png", ".webp", ".heic")

# POS 收銀台的分類比主檔粗——現場靠手機操作，tab 太多要一直橫滑。
# 小何 2026-07-31 決定：凸版兩類合併、海報兩類合併、冰箱貼春節併回冰箱貼。
# ⚠️ 改這裡一定要同步改 index.html 的 SEED_CATS、CAT_COLORS、PROMO_RULES 的 cats，
#    並把 PATCH_VER +1。
CAT_MAP = {
    "凸版明信片": "凸版", "凸版方形": "凸版",
    "A4海報": "海報", "細長海報": "海報",
    "冰箱貼-春節": "冰箱貼",
}
CAT_ORDER = ["明信片", "凸版", "海報", "造型卡", "貼紙",
             "書籤", "L夾", "冰箱貼", "帆布袋", "水晶貼"]

def num(v, default=0):
    try:
        if v is None: return default
        if isinstance(v, str):
            s = v.strip()
            if s in ("", "—", "-", "待定"): return default
            return int(float(s))
        return int(v)
    except Exception:
        return default

def build_products():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["文博商品清單"]
    prods = []
    for r in ws.iter_rows(min_row=6, values_only=True):
        code, name, cat, price_raw, bring = r[1], r[2], r[3], r[5], r[9]
        if code is None: continue          # 分類標題列 / 空列
        code = str(code).strip()
        if cat == "水晶貼": continue         # 水晶貼另外合成尺寸 SKU
        stock = num(bring, 0)
        # 定價
        p = num(price_raw, None) if price_raw not in (None, "待定", "—") else None
        if p is None or p == 0:
            if cat == "造型卡": p = 60
            elif cat == "L夾": p = 100
            elif cat == "凸版明信片": p = 80
            elif cat == "凸版方形": p = 250
            elif cat == "冰箱貼-春節": p = 160 if code.startswith("RH") else 150
            elif cat == "書籤": p = 29      # 贈品，但可 $29 加購
            elif cat == "帆布袋": p = 450
            else: p = 0
        # 價格 fallback 用主檔類別判斷完了，才映射成 POS 的粗分類
        prods.append({"id": code, "name": str(name).strip(),
                      "category": CAT_MAP.get(cat, cat),
                      "price": p, "stock": stock, "sold": 0})
    # 水晶貼 3 尺寸 SKU（庫存從盤點表加總）
    wsx = wb["水晶貼款式盤點表"]
    tier = {"小片": [35, 0], "中片": [70, 0], "大片": [120, 0]}
    for r in wsx.iter_rows(min_row=4, values_only=True):
        lvl, unit, bring = r[2], r[3], r[4]
        if lvl in tier and bring: tier[lvl][1] += num(bring, 0)
    prods += [
        {"id": "CS2601-S", "name": "水晶貼（小片）", "category": "水晶貼", "price": tier["小片"][0], "stock": tier["小片"][1], "sold": 0},
        {"id": "CS2601-M", "name": "水晶貼（中片）", "category": "水晶貼", "price": tier["中片"][0], "stock": tier["中片"][1], "sold": 0},
        {"id": "CS2601-L", "name": "水晶貼（大片）", "category": "水晶貼", "price": tier["大片"][0], "stock": tier["大片"][1], "sold": 0},
    ]
    # 依分類順序排序
    prods.sort(key=lambda x: (CAT_ORDER.index(x["category"]) if x["category"] in CAT_ORDER else 99))
    return prods

def to_data_url(path):
    img = Image.open(path)
    try: img = ImageOps.exif_transpose(img)
    except Exception: pass
    if img.mode != "RGB": img = img.convert("RGB")
    w, h = img.size
    if max(w, h) > MAX_SIDE:
        if w >= h: img = img.resize((MAX_SIDE, round(h*MAX_SIDE/w)), Image.LANCZOS)
        else: img = img.resize((round(w*MAX_SIDE/h), MAX_SIDE), Image.LANCZOS)
    buf = io.BytesIO(); img.save(buf, format="JPEG", quality=QUALITY, optimize=True)
    return "data:image/jpeg;base64," + base64.b64encode(buf.getvalue()).decode()

# 這些貨號刻意不附圖（現有的是會誤導的佔位圖，寧可留空白卡）
SKIP_IMG_IDS = {"CS2601-S", "CS2601-M", "CS2601-L"}

def attach_images(prods):
    files = [f for f in os.listdir(IMG_DIR) if f.lower().endswith(EXTS)]
    matched, missing, kb = 0, [], 0
    for p in prods:
        pid = p["id"]
        if pid in SKIP_IMG_IDS: continue
        # 水晶貼三尺寸共用 CS2601 那張
        match_key = "CS2601" if pid.startswith("CS2601") else pid
        cands = [f for f in files if os.path.splitext(f)[0] == match_key] or \
                [f for f in files if f.startswith(match_key)]
        if not cands:
            missing.append(f"{pid} {p['name']}（{p['category']}）"); continue
        url = to_data_url(os.path.join(IMG_DIR, sorted(cands)[0]))
        p["img"] = url; kb += len(url)/1024; matched += 1
    return matched, missing, kb

def main():
    prods = build_products()
    matched, missing, kb = attach_images(prods)
    json.dump({"products": prods}, open(OUT_JSON, "w", encoding="utf-8"),
              ensure_ascii=False, indent=2)
    # seed snippet（無圖）
    lines = []
    for p in prods:
        nm = p["name"].replace('"', '\\"')
        lines.append(f'      {{id:"{p["id"]}",name:"{nm}",category:"{p["category"]}",price:{p["price"]},stock:{p["stock"]},sold:0}},')
    open(OUT_SEED, "w", encoding="utf-8").write("\n".join(lines))

    from collections import Counter
    print(f"✅ {len(prods)} 款；含圖 {matched}，總大小約 {kb/1024:.2f} MB")
    print("分類分布：", dict(Counter(p["category"] for p in prods)))
    if missing:
        print(f"\n⚠️ 這 {len(missing)} 款沒有照片（之後拍好丟進商品圖片/、檔名以貨號開頭，再跑一次）：")
        for m in missing: print("   -", m)

if __name__ == "__main__":
    main()
